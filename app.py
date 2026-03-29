"""
AngelOne SmartAPI - Candle Data Fetcher
Flask Web Application

Deploy on any server (Ubuntu VPS, Render, Railway, etc.)

Install:
    pip install -r requirements.txt

Run:
    python app.py
"""

import io
import pyotp
import pandas as pd
from flask import Flask, render_template, request, send_file, jsonify
from SmartApi import SmartConnect
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)


# ──────────────────────────────────────────────
#  AngelOne API helpers
# ──────────────────────────────────────────────

def angel_login(api_key, client_id, password, totp_secret):
    obj = SmartConnect(api_key=api_key)
    totp = pyotp.TOTP(totp_secret).now()
    data = obj.generateSession(client_id, password, totp)
    if not data.get("status"):
        raise ValueError(f"Login failed: {data.get('message', 'Unknown error')}")
    return obj


def fetch_candles(session, symbol_token, exchange, interval, start_dt, end_dt):
    start = datetime.strptime(start_dt, "%Y-%m-%d %H:%M")
    end   = datetime.strptime(end_dt,   "%Y-%m-%d %H:%M")

    all_candles = []
    chunk_start = start

    # AngelOne limits intraday requests to ~60 days per call
    chunk_map = {
        "ONE_MINUTE":     30,
        "THREE_MINUTE":   30,
        "FIVE_MINUTE":    60,
        "TEN_MINUTE":     60,
        "FIFTEEN_MINUTE": 60,
        "THIRTY_MINUTE":  60,
        "ONE_HOUR":       200,
        "ONE_DAY":        2000,
    }
    chunk_days = chunk_map.get(interval, 60)

    while chunk_start < end:
        chunk_end = min(chunk_start + timedelta(days=chunk_days), end)
        params = {
            "exchange":    exchange,
            "symboltoken": symbol_token,
            "interval":    interval,
            "fromdate":    chunk_start.strftime("%Y-%m-%d %H:%M"),
            "todate":      chunk_end.strftime("%Y-%m-%d %H:%M"),
        }
        resp = session.getCandleData(params)
        if resp.get("status") and resp.get("data"):
            all_candles.extend(resp["data"])
        chunk_start = chunk_end + timedelta(minutes=1)

    return all_candles


def build_dataframe(candles):
    df = pd.DataFrame(candles, columns=["Timestamp", "Open", "High", "Low", "Close", "Volume"])
    df["Timestamp"] = pd.to_datetime(df["Timestamp"])
    df = df.sort_values("Timestamp").reset_index(drop=True)
    for col in ["Open", "High", "Low", "Close"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").round(2)
    df["Volume"] = pd.to_numeric(df["Volume"], errors="coerce").astype("Int64")
    return df


def create_excel(df, symbol, exchange, interval, start_dt, end_dt):
    wb = Workbook()
    ws = wb.active
    ws.title = "Candles"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="C0392B")
    alt_fill    = PatternFill("solid", start_color="FDEDEC")
    center      = Alignment(horizontal="center", vertical="center")
    thin_border = Border(bottom=Side(style="thin", color="E0E0E0"))
    cols        = ["Timestamp", "Open", "High", "Low", "Close", "Volume"]

    # Title row
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = f"{symbol} ({exchange})  |  {interval.replace('_',' ')} Candles  |  {start_dt}  →  {end_dt}"
    t.font      = Font(name="Calibri", bold=True, size=12, color="C0392B")
    t.alignment = center
    ws.row_dimensions[1].height = 22

    # Header
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font = header_font; c.fill = header_fill; c.alignment = center
    ws.row_dimensions[2].height = 18

    # Data rows
    for ri, row in enumerate(df.itertuples(index=False), 3):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Calibri", size=9)
            c.alignment = center
            c.border = thin_border
            if fill: c.fill = fill
            col_name = cols[ci - 1]
            if col_name == "Timestamp":  c.number_format = "DD-MMM-YYYY HH:MM"
            elif col_name == "Volume":   c.number_format = "#,##0"
            else:                        c.number_format = "#,##0.00"

    widths = {"Timestamp": 22, "Open": 12, "High": 12, "Low": 12, "Close": 12, "Volume": 14}
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths[col]
    ws.freeze_panes = "A3"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    last_row = len(df) + 2
    lbl_fill = PatternFill("solid", start_color="FADBD8")
    summary = [
        ("Symbol",        symbol),
        ("Exchange",      exchange),
        ("Interval",      interval.replace("_", " ")),
        ("Start Date",    start_dt),
        ("End Date",      end_dt),
        ("Total Candles", len(df)),
        ("Highest High",  f"=MAX(Candles!C3:C{last_row})"),
        ("Lowest Low",    f"=MIN(Candles!D3:D{last_row})"),
        ("Avg Close",     f"=AVERAGE(Candles!E3:E{last_row})"),
        ("Total Volume",  f"=SUM(Candles!F3:F{last_row})"),
        ("Avg Volume",    f"=AVERAGE(Candles!F3:F{last_row})"),
    ]
    ws2.merge_cells("A1:B1")
    h = ws2["A1"]
    h.value = "Summary"
    h.font = Font(name="Calibri", bold=True, size=13, color="C0392B")
    h.alignment = center
    for i, (label, val) in enumerate(summary, 2):
        lc = ws2.cell(row=i, column=1, value=label)
        vc = ws2.cell(row=i, column=2, value=val)
        lc.font = Font(name="Calibri", bold=True, size=10); lc.fill = lbl_fill
        vc.font = Font(name="Calibri", size=10)
        vc.alignment = Alignment(horizontal="left")
        if label in ("Highest High", "Lowest Low", "Avg Close"): vc.number_format = "#,##0.00"
        elif label in ("Total Volume", "Avg Volume"):            vc.number_format = "#,##0"
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────
#  Routes
# ──────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/fetch", methods=["POST"])
def fetch():
    try:
        data = request.get_json()

        # Validate required fields
        required = ["api_key", "client_id", "password", "totp_secret",
                    "symbol", "symbol_token", "exchange", "interval",
                    "start_date", "end_date"]
        for field in required:
            if not data.get(field, "").strip():
                return jsonify({"success": False, "error": f"Missing field: {field}"}), 400

        # Login
        session = angel_login(
            data["api_key"], data["client_id"],
            data["password"], data["totp_secret"]
        )

        # Fetch candles
        start_dt = f"{data['start_date']} {data.get('start_time', '09:15')}"
        end_dt   = f"{data['end_date']} {data.get('end_time', '15:30')}"

        raw = fetch_candles(
            session, data["symbol_token"], data["exchange"],
            data["interval"], start_dt, end_dt
        )

        if not raw:
            return jsonify({"success": False, "error": "No data returned. Check symbol token, date range, and ensure it's within market hours (09:15–15:30 IST)."}), 404

        df  = build_dataframe(raw)
        buf = create_excel(df, data["symbol"], data["exchange"],
                           data["interval"], start_dt, end_dt)

        filename = f"{data['symbol']}_{data['interval']}_{data['start_date']}_to_{data['end_date']}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 401
    except Exception as e:
        return jsonify({"success": False, "error": f"Server error: {str(e)}"}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=4000, debug=False)
